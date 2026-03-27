import io
import os
import boto3
from botocore.client import Config
import pandas as pd
from openpyxl import load_workbook
from google import genai
from dotenv import load_dotenv
import time

load_dotenv()

API_KEY = os.getenv("API_KEY")

AWS_BDD_INPUT_BUCKET = os.getenv("aws_bdd_input_bucket")
AWS_BDD_OUTPUT_BUCKET = os.getenv("aws_bdd_output_bucket")
AWS_ARCHIVE_BUCKET = os.getenv("aws_bdd_archive_bucket")
AWS_ACCESS_KEY_ID = os.getenv("aws_access_key_id")
AWS_SECRET_ACCESS_KEY = os.getenv("aws_secret_access_key")
AWS_LOB_FILES = os.getenv("aws_lob_files")
AWS_TEST_OUTPUT_BUCKET = os.getenv("aws_test_output_bucket")
AWS_REGION = os.getenv("AWS_REGION")

# Gemini client
client = genai.Client(api_key=API_KEY)

MODEL_NAME = "models/gemini-2.5-flash"


# S3 Client
s3_client = boto3.client(
    "s3",
    region_name=AWS_REGION,
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
    config=Config(signature_version="s3v4")
)


def generate_llm_response(prompt: str):
    """Reusable Gemini call"""
    try:
        response = client.models.generate_content(
            model=MODEL_NAME,
            contents=prompt
        )
        return response.text
    except Exception as e:
        print("LLM Error:", e)
        return None


def generate_download_url(bucket, key, expiry=3600):
    """Generate secure download URL for S3 object"""
    try:
        url = s3_client.generate_presigned_url(
            "get_object",
            Params={
                "Bucket": bucket,
                "Key": key
            },
            ExpiresIn=expiry
        )
        return url
    except Exception as e:
        print("Error generating download URL:", e)
        return None

def upload_file_to_s3(username):
    try:
        file = f"./static/uploads/{username}_input.xlsx"
        s3_client.upload_file(file, AWS_BDD_INPUT_BUCKET, f"{username}_input.xlsx")
        return True
    except Exception as e:
        print(e)
        return False


def generate_bdd_from_jira(user_story):

    responses = []

    for story in user_story:

        prompt = f"""
        You are a QA expert. Generate a BDD feature file for the following user story.

        User Story: {story}

        Rules:
        - Use strict Gherkin syntax: Feature, Scenario, Given, When, Then, And
        - Generate 2-3 Scenarios covering happy path, edge case, and failure case
        - Add relevant tags like @regression, @smoke, @negative
        - Do NOT include markdown formatting, code blocks, or backticks
        - Do NOT add any explanation or comments outside the feature file
        - Output only raw Gherkin text

        Example format:
        Feature: <feature title>

        @smoke
        Scenario: <scenario title>
            Given <precondition>
            When <action>
            Then <expected result>

        @negative
        Scenario: <negative scenario title>
            Given <precondition>
            When <invalid action>
            Then <expected failure result>
        """

        response = generate_llm_response(prompt)

        responses.append([story, response])

    df1 = pd.DataFrame(responses)

    with io.StringIO() as csv_buffer:

        df1.to_csv(csv_buffer, index=False)

        ts = str(int(time.time()))

        response = s3_client.put_object(
            Bucket=AWS_BDD_OUTPUT_BUCKET,
            Key=f"output_{ts}.csv",
            Body=csv_buffer.getvalue()
        )

        status = response.get("ResponseMetadata", {}).get("HTTPStatusCode")

        key = f"output_{ts}.csv"

        url = generate_download_url(AWS_BDD_OUTPUT_BUCKET, key)

        return url if status == 200 else None


def generate_bdd_scenario(username):

    s3_client_data = s3_client.get_object(
        Bucket=AWS_BDD_INPUT_BUCKET,
        Key=f"{username}_input.xlsx"
    )

    contents = s3_client_data["Body"].read()

    wb = load_workbook(filename=(io.BytesIO(contents)), data_only=True)

    sheet = wb.active

    responses = []

    for row in range(2, sheet.max_row + 1):

        prompt = sheet.cell(row, 1).value

        if not prompt:
            continue

        full_prompt = f"""
Generate BDD scenario in Gherkin feature file format.

User Story:
{prompt}

Return only valid Gherkin format.
"""

        response = generate_llm_response(full_prompt)

        responses.append(response)

    df1 = pd.DataFrame(responses)

    with io.StringIO() as csv_buffer:

        df1.to_csv(csv_buffer, index=False)

        ts = str(int(time.time()))

        response = s3_client.put_object(
            Bucket=AWS_BDD_OUTPUT_BUCKET,
            Key=f"output_{ts}.csv",
            Body=csv_buffer.getvalue()
        )

        status = response.get("ResponseMetadata", {}).get("HTTPStatusCode")

        # Archive original input
        copy_source = {
        'Bucket': AWS_BDD_INPUT_BUCKET,
        'Key': f"{username}_input.xlsx"
        }

        s3_client.copy_object(
            CopySource=copy_source,
            Bucket=AWS_ARCHIVE_BUCKET,
            Key=f"{username}_input_{ts}.xlsx"
        )

        s3_client.delete_object(
            Bucket=AWS_BDD_INPUT_BUCKET,
            Key=f"{username}_input.xlsx"
        )

        key = f"output_{ts}.csv"

        url = generate_download_url(AWS_BDD_OUTPUT_BUCKET, key)

        return url if status == 200 else None


def generate_test_data(lob, state, no_of_test_cases):

    s3_client_data = s3_client.get_object(
        Bucket=AWS_LOB_FILES,
        Key=f"{lob}.txt"
    )

    contents = s3_client_data["Body"].read().decode("utf-8")

    responses = []

    round_of_test_data = int(no_of_test_cases) // 10

    for i in range(round_of_test_data):

        prompt = f"""
Generate 10 test data rows for a {lob} policy.

Constraints:
- state = {state}
- line of business = {lob}

Reference data:
{contents}

Return ONLY CSV format.
"""

        response = generate_llm_response(prompt)

        if i == 0:
            responses.append(response)
        else:
            responses.append("\n".join(response.split("\n")[1:]))

    responses_bytes = ("\n".join(responses)).encode("utf-8")

    ts = str(int(time.time()))

    response = s3_client.put_object(
        Bucket=AWS_TEST_OUTPUT_BUCKET,
        Key=f"{lob}_{ts}.csv",
        Body=responses_bytes
    )

    status = response.get("ResponseMetadata", {}).get("HTTPStatusCode")

    key = f"{lob}_{ts}.csv"

    url = generate_download_url(AWS_TEST_OUTPUT_BUCKET, key)

    return url if status == 200 else None